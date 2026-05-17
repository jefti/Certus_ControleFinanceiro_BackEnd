"""
Exemplo de geração de Excel com dados fictícios.
Este script demonstra como o dashboard export funciona com dados de exemplo.
Versão simplificada que não depende do main.py para evitar problemas de dependências.
"""

from datetime import date
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Cores
_RED_BG       = "FFCCCC"   # light red  – expense row fill
_GREEN_BG     = "CCFFCC"   # light green – income row fill
_RED_DARK     = "C00000"   # dark red   – section header (expenses side)
_GREEN_DARK   = "375623"   # dark green – section header (income side)
_BLUE_DARK    = "1F4E79"   # dark blue  – main header row
_GREY_GROUP   = "D9D9D9"   # grey       – cost-centre group header
_ORANGE       = "F4B942"   # orange     – subtotal row
_WHITE        = "FFFFFF"

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_COLS = ["Descrição", "Tipo", "Status", "Data Vencimento", "Data Pagamento", "Valor (R$)"]
_VAL_COL = 6   # 1-indexed column for "Valor (R$)"
_BRL_FMT = 'R$ #,##0.00'


def _cell_style(
    cell,
    *,
    bold: bool = False,
    color: str = "000000",
    bg: str | None = None,
    align: str = "left",
    border: bool = False,
    number_format: str | None = None,
) -> None:
    cell.font = Font(bold=bold, color=color)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=False)
    if border:
        cell.border = _THIN_BORDER
    if number_format:
        cell.number_format = number_format


def _auto_column_widths(ws) -> None:
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0) for c in col_cells
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 55)


def _status_pagamento(titulo: dict) -> str:
    if titulo.get("dataPagamento"):
        return "Pago"
    venc = titulo.get("dataVencimento")
    if venc:
        try:
            if date.fromisoformat(str(venc)) < date.today():
                return "Vencido"
        except ValueError:
            pass
    return "Em Aberto"


def _group_by_cost_centre(titulos: list[dict]) -> dict[str, list[dict]]:
    """Return an ordered dict: cost-centre name → list of titulos."""
    groups: dict[str, list[dict]] = {}
    for t in titulos:
        centros = t.get("centrosDeCusto") or []
        if not centros:
            groups.setdefault("Sem Centro de Custo", []).append(t)
        else:
            for c in centros:
                groups.setdefault(c.get("descricao") or "Sem Descrição", []).append(t)
    return groups


def _write_group_header(ws, row: int, name: str) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(_COLS))
    _cell_style(ws.cell(row=row, column=1, value=f"  {name.upper()}"),
                bold=True, bg=_GREY_GROUP, align="left", border=True)


def _write_titulo_row(ws, row: int, titulo: dict) -> tuple[float, bool]:
    """Write one titulo row. Returns (valor, is_receber)."""
    is_receber = titulo.get("tipo") == "RECEBER"
    valor = float(titulo.get("valor") or 0)
    bg = _GREEN_BG if is_receber else _RED_BG
    values = [
        titulo.get("descricao"),
        "A Receber" if is_receber else "A Pagar",
        _status_pagamento(titulo),
        titulo.get("dataVencimento"),
        titulo.get("dataPagamento"),
        valor,
    ]
    for col_idx, val in enumerate(values, start=1):
        fmt = _BRL_FMT if col_idx == _VAL_COL else None
        _cell_style(ws.cell(row=row, column=col_idx, value=val), bg=bg, border=True, number_format=fmt)
    return valor, is_receber


def _write_subtotal_row(ws, row: int, saldo: float) -> None:
    _cell_style(ws.cell(row=row, column=_VAL_COL - 1, value="Subtotal do grupo"),
                bold=True, bg=_ORANGE, align="right", border=True)
    _cell_style(ws.cell(row=row, column=_VAL_COL, value=saldo),
                bold=True, bg=_GREEN_BG if saldo >= 0 else _RED_BG, border=True, number_format=_BRL_FMT)


def _write_balance_row(ws, row: int, label: str, value: float,
                       label_bg: str, val_bg: str, font_size: int = 11) -> None:
    lbl = ws.cell(row=row, column=_VAL_COL - 1, value=label)
    val = ws.cell(row=row, column=_VAL_COL, value=value)
    _cell_style(lbl, bold=True, color=_WHITE, bg=label_bg, align="right", border=True)
    _cell_style(val, bold=True, bg=val_bg, border=True, number_format=_BRL_FMT)
    lbl.font = Font(bold=True, color=_WHITE, size=font_size)
    val.font = Font(bold=True, size=font_size)


def _write_extrato_sheet(wb: Workbook, titulos: list[dict]) -> None:
    ws = wb.active
    ws.title = "Extrato"
    ws.row_dimensions[1].height = 22

    for col_idx, col_name in enumerate(_COLS, start=1):
        _cell_style(ws.cell(row=1, column=col_idx, value=col_name),
                    bold=True, color=_WHITE, bg=_BLUE_DARK, align="center", border=True)
    ws.freeze_panes = "A2"

    row = 2
    total_receber = 0.0
    total_pagar   = 0.0

    for group_name, group_titulos in _group_by_cost_centre(titulos).items():
        _write_group_header(ws, row, group_name)
        row += 1

        group_receber = 0.0
        group_pagar   = 0.0
        for titulo in group_titulos:
            valor, is_receber = _write_titulo_row(ws, row, titulo)
            if is_receber:
                group_receber += valor
                total_receber += valor
            else:
                group_pagar += valor
                total_pagar += valor
            row += 1

        _write_subtotal_row(ws, row, group_receber - group_pagar)
        row += 3  # subtotal + 2 blank lines

    saldo_final = total_receber - total_pagar
    positive    = saldo_final >= 0

    _write_balance_row(ws, row,     "TOTAL A RECEBER (+)", total_receber, _GREEN_DARK, _GREEN_BG)
    _write_balance_row(ws, row + 1, "TOTAL A PAGAR (−)",   total_pagar,   _RED_DARK,   _RED_BG)
    _write_balance_row(ws, row + 2, "SALDO FINAL",         saldo_final,
                       _GREEN_DARK if positive else _RED_DARK,
                       _GREEN_BG   if positive else _RED_BG,
                       font_size=12)

    _auto_column_widths(ws)


def _write_resumo_sheet(wb: Workbook, titulos: list[dict]) -> None:
    ws = wb.create_sheet("Resumo")

    total_receber = sum(float(t.get("valor") or 0) for t in titulos if t.get("tipo") == "RECEBER")
    total_pagar   = sum(float(t.get("valor") or 0) for t in titulos if t.get("tipo") == "PAGAR")
    saldo         = total_receber - total_pagar

    statuses = [_status_pagamento(t) for t in titulos]

    rows = [
        ("Data de exportação",       date.today().isoformat(),                    None),
        ("Total de títulos",         len(titulos),                                None),
        ("Títulos a Pagar",          sum(1 for t in titulos if t.get("tipo") == "PAGAR"),  None),
        ("Títulos a Receber",        sum(1 for t in titulos if t.get("tipo") == "RECEBER"), None),
        ("Títulos Pagos",            statuses.count("Pago"),                      None),
        ("Títulos Em Aberto",        statuses.count("Em Aberto"),                 None),
        ("Títulos Vencidos",         statuses.count("Vencido"),                   None),
        ("Total a Receber",          total_receber,                               _BRL_FMT),
        ("Total a Pagar",            total_pagar,                                 _BRL_FMT),
        ("Saldo Final",              saldo,                                       _BRL_FMT),
    ]

    for col_idx, h in enumerate(["Métrica", "Valor"], start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _cell_style(cell, bold=True, color=_WHITE, bg=_BLUE_DARK, align="center", border=True)

    for r_idx, (label, value, fmt) in enumerate(rows, start=2):
        lbl = ws.cell(row=r_idx, column=1, value=label)
        val = ws.cell(row=r_idx, column=2, value=value)
        _cell_style(lbl, border=True)
        _cell_style(val, border=True, number_format=fmt or "@")

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.freeze_panes = "A2"


def _build_excel(titulos: list[dict]) -> bytes:
    wb = Workbook()
    _write_extrato_sheet(wb, titulos)
    _write_resumo_sheet(wb, titulos)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# Dados de exemplo
exemplo_titulos = [
    # Títulos PAGAR - Vendas
    {
        "id": 1,
        "descricao": "Aluguel do escritório - Vendas",
        "tipo": "PAGAR",
        "valor": 3500.00,
        "dataVencimento": "2025-06-10",
        "dataPagamento": "2025-06-09",
        "centrosDeCusto": [{"id": 1, "descricao": "Vendas"}]
    },
    {
        "id": 2,
        "descricao": "Internet e telefonia - Vendas",
        "tipo": "PAGAR",
        "valor": 250.00,
        "dataVencimento": "2025-06-15",
        "dataPagamento": None,
        "centrosDeCusto": [{"id": 1, "descricao": "Vendas"}]
    },
    {
        "id": 3,
        "descricao": "Salários - Vendas",
        "tipo": "PAGAR",
        "valor": 15000.00,
        "dataVencimento": "2025-06-30",
        "dataPagamento": None,
        "centrosDeCusto": [{"id": 1, "descricao": "Vendas"}]
    },
    
    # Títulos RECEBER - Vendas
    {
        "id": 4,
        "descricao": "Venda de Produtos - Fatura #001",
        "tipo": "RECEBER",
        "valor": 8500.00,
        "dataVencimento": "2025-06-20",
        "dataPagamento": "2025-06-18",
        "centrosDeCusto": [{"id": 1, "descricao": "Vendas"}]
    },
    {
        "id": 5,
        "descricao": "Consultoria - Fatura #002",
        "tipo": "RECEBER",
        "valor": 2000.00,
        "dataVencimento": "2025-06-25",
        "dataPagamento": None,
        "centrosDeCusto": [{"id": 1, "descricao": "Vendas"}]
    },
    
    # Títulos PAGAR - Administrativo
    {
        "id": 6,
        "descricao": "Energia elétrica",
        "tipo": "PAGAR",
        "valor": 800.00,
        "dataVencimento": "2025-06-10",
        "dataPagamento": "2025-06-10",
        "centrosDeCusto": [{"id": 2, "descricao": "Administrativo"}]
    },
    {
        "id": 7,
        "descricao": "Material de escritório",
        "tipo": "PAGAR",
        "valor": 450.00,
        "dataVencimento": "2025-05-20",  # Vencido
        "dataPagamento": None,
        "centrosDeCusto": [{"id": 2, "descricao": "Administrativo"}]
    },
    {
        "id": 8,
        "descricao": "Software de gestão",
        "tipo": "PAGAR",
        "valor": 300.00,
        "dataVencimento": "2025-06-05",
        "dataPagamento": None,
        "centrosDeCusto": [{"id": 2, "descricao": "Administrativo"}]
    },
    
    # Títulos RECEBER - Administrativo
    {
        "id": 9,
        "descricao": "Devolução de depósito",
        "tipo": "RECEBER",
        "valor": 1200.00,
        "dataVencimento": "2025-06-01",
        "dataPagamento": "2025-06-01",
        "centrosDeCusto": [{"id": 2, "descricao": "Administrativo"}]
    },
    
    # Títulos PAGAR - RH
    {
        "id": 10,
        "descricao": "Treinamento de funcionários",
        "tipo": "PAGAR",
        "valor": 2000.00,
        "dataVencimento": "2025-06-30",
        "dataPagamento": None,
        "centrosDeCusto": [{"id": 3, "descricao": "RH"}]
    },
    {
        "id": 11,
        "descricao": "Vale alimentação",
        "tipo": "PAGAR",
        "valor": 3500.00,
        "dataVencimento": "2025-06-30",
        "dataPagamento": None,
        "centrosDeCusto": [{"id": 3, "descricao": "RH"}]
    },
    
    # Título sem centro de custo
    {
        "id": 12,
        "descricao": "Despesa diversa",
        "tipo": "PAGAR",
        "valor": 500.00,
        "dataVencimento": "2025-06-15",
        "dataPagamento": None,
        "centrosDeCusto": []
    },
]


def main():
    """Gera um arquivo Excel de exemplo."""
    # Gera o Excel
    excel_bytes = _build_excel(exemplo_titulos)
    
    # Salva o arquivo
    filename = f"exemplo_financeiro_{date.today().isoformat()}.xlsx"
    output_path = os.path.join(os.path.dirname(__file__), filename)
    
    with open(output_path, "wb") as f:
        f.write(excel_bytes)
    
    print(f"✓ Arquivo gerado com sucesso: {filename}")
    print(f"  Tamanho: {len(excel_bytes):,} bytes")
    print(f"\nDados incluídos:")
    print(f"  - Total de títulos: {len(exemplo_titulos)}")
    print(f"  - Títulos PAGAR: {sum(1 for t in exemplo_titulos if t['tipo'] == 'PAGAR')}")
    print(f"  - Títulos RECEBER: {sum(1 for t in exemplo_titulos if t['tipo'] == 'RECEBER')}")
    
    total_pagar = sum(t["valor"] for t in exemplo_titulos if t["tipo"] == "PAGAR")
    total_receber = sum(t["valor"] for t in exemplo_titulos if t["tipo"] == "RECEBER")
    saldo = total_receber - total_pagar
    
    print(f"\nResumo financeiro:")
    print(f"  - Total a Receber: R$ {total_receber:,.2f}")
    print(f"  - Total a Pagar: R$ {total_pagar:,.2f}")
    print(f"  - Saldo Final: R$ {saldo:,.2f}")


if __name__ == "__main__":
    main()
